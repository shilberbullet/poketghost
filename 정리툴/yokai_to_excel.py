# 사용자가 새로 추가한 텍스트를 포함한 전체 요괴 데이터로 다시 엑셀 생성

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 텍스트 데이터
extended_text = '''
 
# 일반 요괴
도깨비,MONSTER,200,150,180,170,밤에 나타나 사람들을 놀라게 하지만 가끔씩 도움을 주기도 한다.,불꽃치기;지옥불;맹화;가시돌진;사념의박치기;강철꼬리치기;몸통박치기;번개
구미호,ANIMAL,180,160,190,200,아홉 개의 꼬리를 가진 여우 요괴. 사람의 모습으로 변신하여 사람들을 유혹한다.,꼬리치기;유혹하기;물기;깨물어부수기;마구 할퀴기;맹독침;물어뜯기;짖기
귀신,GHOST,190,140,160,220,원한을 풀지 못한 영혼이 변한 요괴. 매우 빠른 속도로 움직인다.,음파;염력;얼음창;바람칼;번개;흙폭탄;불꽃치기;돌던지기
여우,ANIMAL,160,150,170,230,꼬리가 여러 개 달린 여우 요괴. 민첩한 움직임이 특징이다.,꼬리치기;유혹하기;물기;깨물어부수기;마구 할퀴기;맹독침;물어뜯기;짖기
산신령,HUMAN,170,200,200,160,산을 지키는 신령. 강한 방어력과 체력을 가지고 있다.,돌던지기;바람칼;흙폭탄;후려치기;분노의주먹;얼음창;번개;불꽃치기

# 보스 요괴
도깨비불,EVIL_SPIRIT,230,180,220,190,도깨비의 왕. 강력한 불꽃을 다루는 능력을 가지고 있다.,지옥불;맹화;불꽃치기;가시돌진;사념의박치기;강철꼬리치기;몸통박치기;번개
구미호왕,ANIMAL,210,190,230,220,구미호의 왕. 아홉 개의 꼬리로 강력한 공격을 한다.,꼬리치기;유혹하기;물기;깨물어부수기;마구 할퀴기;맹독침;물어뜯기;짖기
귀신왕,GHOST,220,170,200,240,귀신의 왕. 매우 빠른 속도로 적을 공격한다.,음파;염력;얼음창;바람칼;번개;흙폭탄;불꽃치기;돌던지기
산신령왕,HUMAN,200,230,240,190,산신령의 왕. 강력한 방어력으로 적의 공격을 막아낸다.,돌던지기;바람칼;흙폭탄;후려치기;분노의주먹;얼음창;번개;불꽃치기
여우왕,ANIMAL,190,200,220,250,여우의 왕. 매우 빠른 속도로 적을 공격한다.,꼬리치기;유혹하기;물기;깨물어부수기;마구 할퀴기;맹독침;물어뜯기;짖기 

# 패러독스 요괴
이계의 망령,EVIL_SPIRIT,250,180,210,200,이계에서 온 정체불명의 망령. 현실과의 경계가 모호하다.,혼돈의파동;차원의칼날;어둠의손길;영혼흡수
차원 파수꾼,MONSTER,240,170,220,210,차원을 지키는 수호자. 강력한 힘을 자랑한다.,차원방패;파멸의주먹;공간왜곡;수호의외침
심연의 그림자,GHOST,230,160,200,220,심연에서 태어난 그림자. 모든 빛을 흡수한다.,그림자베기;암흑폭풍;속삭임;절망의눈빛
운명의 조각자,HUMAN,220,200,230,190,운명을 조각하는 자. 미래를 예지한다.,운명베기;예지타격;시간왜곡;심판의칼날
시간의 파편,EVIL_SPIRIT,210,150,190,230,시간의 흐름에서 떨어져 나온 존재.,시간단절;과거회상;미래도약;영원의속삭임
공허의 사도,MONSTER,200,180,210,210,공허를 따르는 사도. 감정을 느끼지 않는다.,공허의창;무감정공격;침묵의포효;무의의손길
붕괴의 인도자,GHOST,220,170,220,200,붕괴를 이끄는 자. 모든 것을 무로 돌린다.,붕괴의손짓;파멸의속삭임;무의식공격;종말의울음
경계의 방랑자,HUMAN,210,160,200,220,경계를 떠도는 방랑자. 어디에도 속하지 않는다.,경계이동;방랑자의칼날;이방인의눈;자유의외침
환영의 군주,ANIMAL,230,180,210,210,환영을 다루는 군주. 실체가 불분명하다.,환영공격;실체변화;환각의포효;허상의손길

# 최종보스
차원의 군주,EVIL_SPIRIT,300,250,300,280,차원의 균열을 지배하는 절대적인 존재. 모든 차원의 힘을 다룬다.,차원파괴;시간지배;공간지배;혼돈의파동;차원의칼날;어둠의손길;영혼흡수;파멸의주먹 

'''

# 기술 데이터
skill_data = '''
# 기술 데이터 파일
# 형식: 이름,기술의 상성,공격력,명중률,PP,설명

# 초급 기술
불꽃치기,MONSTER,30,90,20,불꽃으로 적을 공격한다
돌던지기,HUMAN,25,95,25,돌을 던져 적을 공격한다
꼬리치기,ANIMAL,20,100,25,꼬리로 적을 공격한다
물기,ANIMAL,25,90,20,이빨로 적을 물어뜯는다
할퀴기,ANIMAL,25,90,20,날카로운 발톱으로 할퀴어 공격한다
몸통박치기,MONSTER,30,85,20,몸 전체로 부딪혀 공격한다
독침,ANIMAL,20,95,30,독침을 쏘아 공격한다
짖기,ANIMAL,15,100,30,큰 소리로 적의 명중률을 낮춘다
음파,GHOST,25,90,25,강한 음파로 정신 피해를 입힌다
유혹하기,ANIMAL,15,85,30,적을 유혹하여 공격력을 낮춘다
날개베기,ANIMAL,30,90,15,날개를 이용해 회전 베기를 가한다
속임수베기,HUMAN,35,100,10,기습 공격으로반드시 먼저 공격한다
박치기,MONSTER,30,95,15,연속으로 머리로 들이받는다

# 중급 기술
바람칼,HUMAN,35,85,20,바람으로 만든 칼날을 날린다
얼음창,GHOST,40,80,15,얼음으로 만든 창을 던진다
물어뜯기,ANIMAL,35,80,20,이빨로 적을 물어뜯는다
후려치기,HUMAN,35,80,20,무거운 일격으로 적을 가격한다
맹독침,ANIMAL,35,80,20,강한 독을 가진 독침을 쏘아 공격한다
흙폭탄,HUMAN,30,90,20,흙으로 만든 폭탄을 던진다
마구 할퀴기,ANIMAL,40,85,10,연속으로 할퀴어 강한 피해를 준다
염력,GHOST,35,85,20,염력으로 물건을 던져져 공격한다
울부짖기,ANIMAL,20,100,5,자신의 공격력을 증가시킨다
연속박치기,MONSTER,40,80,/15,연속으로 머리로 들이받는다

# 고급 기술
지옥불,EVIL_SPIRIT,50,70,5,지옥의 불꽃으로 적을 태운다
가시돌진,MONSTER,40,85,15,몸의 가시로 돌진하며 상대에게 큰 피해를 준다
분노의주먹,HUMAN,45,80,10,분노를 담아 강한 한 방을 날린다
사념의박치기,MONSTER,50,80,5,강하게 부딪힌다다
강철꼬리치기,MONSTER,45,80,10,단단한 꼬리로 강하게 내리친다
맹화,MONSTER,50,85,5,강한 불꽃으로 공격헌다
깨물어부수기,ANIMAL,40,90,15,이빨로 깨물어 부순다다
번개,GHOST,45,75,10,번개로 적을 공격한다
철퇴내리치기,HUMAN,50,75,5,무거운 철퇴를 휘둘러 큰 피해를 준다
지옥기운,EVIL_SPIRIT,60,65,5,지옥의 어둠으로 적에게 큰 피해를 준다
이단도약,MONSTER,45,85,10,두 번 튕겨 뛰며 큰 피해를 준다
망령의눈빛,GHOST,100,30,5,행동불능으로 만들 확률 30%

# 패러독스 요괴 기술
혼돈의파동,EVIL_SPIRIT,85,90,10,혼돈의 파동으로 적을 공격
차원의칼날,EVIL_SPIRIT,80,95,10,차원의 칼날로 적을 베어버림
어둠의손길,EVIL_SPIRIT,75,100,15,어둠의 손길로 적을 공격
영혼흡수,EVIL_SPIRIT,70,100,15,적의 영혼을 흡수하여 HP 회복
차원방패,EVIL_SPIRIT,0,100,10,차원의 방패로 자신을 보호
파멸의주먹,EVIL_SPIRIT,85,90,10,파멸을 부르는 강력한 주먹
공간왜곡,EVIL_SPIRIT,0,100,5,공간을 왜곡하여 적의 공격을 무효화
수호의외침,EVIL_SPIRIT,0,100,5,강력한 외침으로 자신의 방어력 증가
그림자베기,GHOST,75,95,10,그림자로 적을 베어버림
암흑폭풍,GHOST,80,90,10,암흑의 폭풍으로 적을 공격
속삭임,GHOST,0,100,5,적의 공격력을 낮추는 속삭임
절망의눈빛,GHOST,0,100,5,적을 절망에 빠뜨리는 눈빛
운명베기,HUMAN,80,90,10,운명을 베어버리는 강력한 공격
예지타격,HUMAN,75,95,10,미래를 예지하여 정확한 타격
시간왜곡,HUMAN,0,100,5,시간을 왜곡하여 적의 행동을 봉쇄
심판의칼날,HUMAN,85,85,10,심판의 칼날로 적을 베어버림
시간단절,EVIL_SPIRIT,0,100,5,시간의 흐름을 끊어버림
과거회상,EVIL_SPIRIT,0,100,5,과거의 기억으로 자신을 강화
미래도약,EVIL_SPIRIT,0,100,5,미래로 도약하여 회피
영원의속삭임,EVIL_SPIRIT,0,100,5,영원한 속삭임으로 적을 약화
공허의창,EVIL_SPIRIT,80,90,10,공허의 창으로 적을 관통
무감정공격,EVIL_SPIRIT,75,95,10,감정 없는 공격으로 적을 공격
침묵의포효,EVIL_SPIRIT,0,100,5,침묵의 포효로 적을 약화
무의의손길,EVIL_SPIRIT,70,100,15,무의의 손길로 적을 공격
붕괴의손짓,GHOST,0,100,5,붕괴의 손짓으로 적을 약화
파멸의속삭임,GHOST,0,100,5,파멸의 속삭임으로 적을 공격
무의식공격,GHOST,75,95,10,무의식의 공격으로 적을 공격
종말의울음,GHOST,0,100,5,종말을 알리는 울음으로 적을 약화
경계이동,HUMAN,0,100,5,경계를 이동하여 회피
방랑자의칼날,HUMAN,80,90,10,방랑자의 칼날로 적을 베어버림
이방인의눈,HUMAN,0,100,5,이방인의 눈으로 적을 약화
자유의외침,HUMAN,0,100,5,자유의 외침으로 자신을 강화
환영공격,ANIMAL,75,95,10,환영으로 적을 공격
실체변화,ANIMAL,0,100,5,실체를 변화시켜 회피
환각의포효,ANIMAL,0,100,5,환각의 포효로 적을 혼란
허상의손길,ANIMAL,70,100,15,허상의 손길로 적을 공격

# 최종보스 기술
차원파괴,EVIL_SPIRIT,100,80,5,차원을 파괴하는 강력한 공격
시간지배,EVIL_SPIRIT,90,85,5,시간을 지배하여 적의 행동을 봉쇄
공간지배,EVIL_SPIRIT,95,80,5,공간을 지배하여 적을 압박

'''

# 기술 타입 매핑 딕셔너리 생성
skill_type_map = {}
for line in skill_data.strip().splitlines():
    parts = line.split(",")
    if len(parts) >= 2:
        skill_name = parts[0].strip()
        skill_type = parts[1].strip()
        skill_type_map[skill_name] = skill_type

def parse_yokai_data(text_data):
    """텍스트 형식의 요괴 데이터를 파싱합니다."""
    records = []
    current_kind = "일반"  # 기본값
    for line in text_data.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith('#'):
            if "일반" in line:
                current_kind = "일반"
            elif "보스" in line:
                current_kind = "보스"
            elif "패러독스" in line:
                current_kind = "패러독스"
            elif "최종보스" in line:
                current_kind = "최종보스"
            continue
        
        parts = line.split(",")
        if len(parts) < 7:
            continue
        
        name = parts[0].strip()
        type_ = parts[1].strip().replace("TYPE_", "")
        atk, df, hp, spd = map(int, parts[2:6])
        description = parts[6].strip()
        skills = ",".join(parts[7:]).replace(";", "; ").strip()
        skill_list = [s.strip() for s in skills.split(";") if s.strip()]
        while len(skill_list) < 10:
            skill_list.append("")
        
        # 종족값 계산
        total_stats = atk + df + hp + spd
        
        records.append([name, type_, atk, df, hp, spd, total_stats, description] + skill_list[:10] + [current_kind])
    
    return pd.DataFrame(records, columns=["이름", "타입", "공격력", "방어력", "체력", "스피드", "종족값", "도감설명", 
                                        "기술1", "기술2", "기술3", "기술4", "기술5", "기술6", "기술7", "기술8", "기술9", "기술10", "요괴종류"])

# 중복 제거 및 데이터 정리
unique_records = []
seen = set()
current_kind = "일반"
for line in extended_text.strip().splitlines():
    line = line.strip()
    if not line:
        continue
    if line.startswith('#'):
        if "일반" in line:
            current_kind = "일반"
        elif "보스" in line:
            current_kind = "보스"
        elif "패러독스" in line:
            current_kind = "패러독스"
        elif "최종보스" in line:
            current_kind = "최종보스"
        continue
    parts = line.split(",")
    if len(parts) < 7:
        continue
    name = parts[0].strip()
    type_ = parts[1].strip().replace("TYPE_", "")
    atk, df, hp, spd = map(int, parts[2:6])
    description = parts[6].strip()
    skills = ",".join(parts[7:]).replace(";", "; ").strip()
    skill_list = [s.strip() for s in skills.split(";") if s.strip()]
    while len(skill_list) < 10:
        skill_list.append("")
    # 종족값 계산
    total_stats = atk + df + hp + spd
    key = (name, type_, atk, df, hp, spd, description, tuple(skill_list))
    if key not in seen:
        seen.add(key)
        unique_records.append([name, type_, atk, df, hp, spd, total_stats, description] + skill_list[:10] + [current_kind])

# DataFrame 생성 (기술1~10 컬럼 추가)
df = pd.DataFrame(unique_records, columns=["이름", "타입", "공격력", "방어력", "체력", "스피드", "종족값", "도감설명", "기술1", "기술2", "기술3", "기술4", "기술5", "기술6", "기술7", "기술8", "기술9", "기술10", "요괴종류"])

# 타입별 색상 지정
type_colors = {
    "EVIL_SPIRIT": "FF9999",
    "GHOST": "CC99FF",
    "MONSTER": "FFFF99",
    "HUMAN": "66CCCC",
    "ANIMAL": "99CC99",
}

# 요괴종류별 색상 지정
yokai_kind_colors = {
    "일반": "FFFFFF",      # 흰색
    "보스": "FFFACD",      # 연노랑
    "패러독스": "E6E6FA",  # 연보라
    "최종보스": "FFB6C1",  # 연빨강
}

# 워크북 구성
wb = Workbook()

# 요괴 도감 시트
ws_yokai = wb.active
ws_yokai.title = "요괴 도감"

# 데이터 입력
for r in dataframe_to_rows(df, index=False, header=True):
    ws_yokai.append(r)

# 테이블 생성
table = Table(displayName="Table1", ref=f"A1:S{len(df)+1}")
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                      showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws_yokai.add_table(table)

# 타입 색상 적용 (타입 컬럼)
for row in ws_yokai.iter_rows(min_row=2, max_row=ws_yokai.max_row, min_col=2, max_col=2):
    for cell in row:
        color = type_colors.get(cell.value, None)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# 기술별 색상 적용 (기술1~10 컬럼)
for col in range(8, 18):  # 기술1~10 컬럼
    for row in ws_yokai.iter_rows(min_row=2, max_row=ws_yokai.max_row, min_col=col, max_col=col):
        for cell in row:
            if cell.value:
                skill_type = skill_type_map.get(cell.value, None)
                if skill_type:
                    color = type_colors.get(skill_type, None)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# 요괴종류 색상 적용 (요괴종류 컬럼)
for row in ws_yokai.iter_rows(min_row=2, max_row=ws_yokai.max_row, min_col=19, max_col=19):
    for cell in row:
        color = yokai_kind_colors.get(cell.value, None)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# 열 너비 자동 조정
for column in ws_yokai.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_yokai.column_dimensions[column[0].column_letter].width = adjusted_width

# 파일 저장
wb.save("요괴_정리.xlsx")
print("엑셀 파일이 요괴_정리.xlsx에 저장되었습니다.")
