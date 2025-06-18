import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

# 첫 번째 기술 데이터
skill_data_1 = '''
# 초급 기술
불꽃치기,MONSTER,25,100,20,불꽃으로 적을 공격한다
돌던지기,HUMAN,25,95,25,돌을 던져 적을 공격한다
꼬리치기,ANIMAL,20,100,25,꼬리로 적을 공격한다
물기,ANIMAL,25,100,20,이빨로 적을 물어뜯는다
할퀴기,ANIMAL,25,100,20,날카로운 발톱으로 할퀴어 공격한다
몸통박치기,MONSTER,25,100,20,몸 전체로 부딪혀 공격한다
독침,ANIMAL,20,100,30,독침을 쏘아 공격한다
짖기,ANIMAL,15,100,30,소리로 적을 공격한다
음파,GHOST,25,90,25,강한 음파로 정신 피해를 입힌다
유혹하기,ANIMAL,15,85,30,적을 유혹해 상대를 혼란스럽게 한다
날개베기,ANIMAL,25,95,25,날개를 이용해 회전 베기를 가한다
망령의눈빛,GHOST,20,100,30,적에게 오싹한 눈빛을 보내 공격한다
정전기,GHOST,20,100,30,정전기를 일으켜격한다
소리지르기,HUMAN,20,100,30,소리를 질러 적을 공격한다
안개,GHOST,25,100,30,안개에 숨어 적을 공격한다
부채베기,HUMAN,25,95,20,부채로 예리하게 베어 공격한다  
진흙던지기,HUMAN,20,100,25,진흙을 뿌려 시야를 방해한다  
속임수,EVIL_SPIRIT,25,95,25,기습적인 공격을 한다
헛소리,EVIL_SPIRIT,20,100,30,정신을 혼미하게 만드는 알 수 없는 소리를 낸다
괴성,GHOST,25,90,25,듣기 싫은 소리를 내 공격한다
불씨던지기,MONSTER,25,90,20,작은 불씨를 던져 화상을 입힌다  
호미찌르기,HUMAN,25,100,20,호미날로 적을 강하게 찌른다  
사슬채찍,ANIMAL,20,100,25,사슬 같은 꼬리로 적을 휘감아 공격한다  
가시털,MONSTER,25,85,20,등의 가시를 세워 적을 찔러 공격한다  
그림자발바닥,GHOST,20,100,20,바닥의 그림자를 타고 빠르게 돌진해 공격한다  
새벽의속삭임,EVIL_SPIRIT,15,85,30,새벽녘의 기운으로 적을 혼란에 빠뜨린다  
바람발,ANIMAL,15,100,30,날개짓으로 찬 바람을 일으켜 적을 얼린다  
독꽃던지기,EVIL_SPIRIT,15,100,30,독이 깃든 꽃잎을 흩뿌려 중독 상태를 유발한다  
돌가루,EVIL_SPIRIT,15,100,30,돌가루를 날려 적의 시야를 가린다  


# 중급 기술
바람칼,HUMAN,35,85,20,바람으로 만든 칼날을 날린다
얼음창,MONSTER,40,80,15,얼음으로 만든 창을 던진다
물어뜯기,ANIMAL,35,85,20,이빨로 적을 물어뜯는다
후려치기,HUMAN,35,85,20,무거운 일격으로 적을 가격한다
맹독침,ANIMAL,35,85,20,강한 독을 가진 독침을 쏘아 공격한다
흙폭탄,HUMAN,30,90,20,흙으로 만든 폭탄을 던진다
마구할퀴기,ANIMAL,40,85,10,연속으로 할퀴어 강한 피해를 준다
염력,GHOST,35,85,20,염력으로 물건을 던져져 공격한다
속임수베기,EVIL_SPIRIT,35,90,10,기습적인 공격을 한다
폭음파,GHOST,35,85,25,강한 폭발을 일으켜 만든 음파로 피해를 입힌다
어퍼컷,HUMAN,35,85,20,강한 어퍼컷을 날펴 피해를 입힌다
울부짖기,ANIMAL,30,80,15,큰 소리를 질러 공격한다
연속박치기,MONSTER,40,80,15,연속으로 머리로 들이받는다
환영술,EVIL_SPIRIT,35,85,15,상대의 두려움을 담은 환영 보여준다
예측공격,HUMAN,30,90,30,상대의 움직임을 파악해서 약점을 공격한다
가시돌진,MONSTER,40,85,15,몸의 가시로 돌진하며 상대에게 큰 피해를 준다
요상한바람,EVIL_SPIRIT,30,90,20,기분 나쁜 바람으로 공격 한다
쇠방울던지기,HUMAN,30,90,25,쇠방울을 던져 타격을 준다  
검은연기,GHOST,30,90,25,검은 연기를 뿜어 정신을 혼미하게 만든다  
망자의손,GHOST,35,85,15,죽은 자의 손이 나타나 덮친다  
달그림자베기,EVIL_SPIRIT,35,90,15,달빛 속 그림자로 벤다  
호랑이눈빛,ANIMAL,35,85,20,호랑이의 기운으로 상대를 얼어붙게 한다  
피눈물베기,EVIL_SPIRIT,40,85,10,증오로 울며 베어내는 원혼의 칼날  
괴성공명,GHOST,40,80,15,귀신의 절규가 공명을 일으켜 타격을 준다  
숨은송곳니,ANIMAL,35,90,15,숨어 있다가 날카로운 이빨로 덮친다  
혼령의숨,GHOST,30,95,20,혼령이 뿜는 숨결로 정신을 흔든다  
번개,MONSTER,40,85,15,번개로 적을 공격한다
꼬리후려치기,ANIMAL,35,90,15,날렵한 꼬리로 적을 후려친다  
불길장벽,MONSTER,35,85,15,불길의 장벽을 세워 적을 태운다  
호랑이발톱,ANIMAL,30,90,90,호랑이의 발톱으로 강하게 찢어낸다  
칼날비,HUMAN,35,80,20,허공에 나타난 칼날을 비처럼 쏟아낸다  
철화살,HUMAN,35,85,20,단단한 철제 화살을 날려 정확히 관통한다  
폭풍의깃,MONSTER,40,80,15,회오리치는 폭풍구름을 만들어 적에게 돌진시킨다  
영혼폭포,GHOST,35,85,15,영혼의 물줄기를 폭포처럼 쏟아내 피해를 준다  
가시벽,ANIMAL,30,90,20,빠르게 솟아나는 가시로 상대를 가둬 공격한다  
달빛무도,EVIL_SPIRIT,35,90,15,달빛 아래 환영 무도를 펼쳐 연속 공격을 가한다  
검은연못,GHOST,30,85,20,어둠의 연못을 소환해 적을 빨아들인다  
귀신의사슬,EVIL_SPIRIT,35,90,15,유령 사슬로 적을 묶어 묵직한 피해를 준다  
무쇠발톱,ANIMAL,35,85,15,단단한 무쇠 같은 발톱으로 적을 내리찍는다  
망각의노래,EVIL_SPIRIT,35,90,15,저승의 멜로디로 적의 기억을 일시적으로 잃게 한다  
환영분열,EVIL_SPIRIT,35,85,20,자신의 환영을 나누어 다수 타격을 가한다  


# 고급 기술
지옥불,EVIL_SPIRIT,50,70,5,지옥의 불꽃으로 적을 태운다
분노의주먹,HUMAN,45,80,10,분노를 담아 강한 한 방을 날린다
사념의박치기,MONSTER,50,80,5,강하게 부딪힌다
강철꼬리치기,ANIMAL,45,80,10,단단한 꼬리로 강하게 내리친다
맹화,MONSTER,50,85,5,강한 불꽃으로 공격헌다
깨물어부수기,ANIMAL,45,85,15,이빨로 깨물어 부순다다
철퇴내리치기,HUMAN,50,75,5,무거운 철퇴를 휘둘러 큰 피해를 준다
지옥기운,EVIL_SPIRIT,50,70,5,지옥의 어둠으로 적에게 큰 피해를 준다
이단도약,MONSTER,45,85,10,두 번 튕겨 뛰며 큰 피해를 준다
무지개발차기,HUMAN,45,80,10,어디서 날라올지 모르는 강력한 발차기를 날린다
지옥기운,EVIL_SPIRIT,50,70,5,지옥의 어둠으로 적에게 큰 피해를 준다
어둠손길,GHOST,45,85,15,상대를 자신이 있는 어둠속으로 끌어들여 공격한다
그림자베기,GHOST,45,85,10,그림자로 적을 베어버린다
영혼흡수,EVIL_SPIRIT,40,80,20,적의 영혼을 흡수한다
혼백폭발,GHOST,50,75,5,혼백의 힘으로 폭발을 일으킨다  
도깨비방망이,MONSTER,50,85,5,도깨비방망이로 내려쳐 적을 공격한다
혼령메아리,GHOST,45,80,10,죽은 혼령들의 메아리로 정신 공격을 가한다
백호의울음,ANIMAL,50,75,5,전설 속 백호의 울음으로 적을 압도한다
무덤기운,GHOST,50,75,5무덤에서 솟아오르는 힘으로 공격한다
뇌우,GHOST,50,75,5,천둥 번개와 비를 내려 공격한다
혼신의부채베기,HUMAN,50,80,5,혼신을 담아 부채로 빠르게 베어낸다
태산벽력,MONSTER,50,75,5,산을 깨뜨리는 벽력으로 일격을 가한다  
달밤유성,GHOST,45,85,10,달빛 속에서 유성이 떨어져 적을 강타한다 
천지개벽,MONSTER,50,75,5,하늘과 땅을 가르는 거대한 충격파를 발사한다  
지옥폭뢰,EVIL_SPIRIT,50,70,10,지옥의 번개를 소환해 전장을 가로지른다  
암흑회오리,GHOST,50,75,5,암흑 에너지의 회오리를 만들어 광역 흡수 피해를 준다  
불사조의노래,MONSTER,45,80,10,불사조의 울음으로 전장을 불태우며 부활 효과를 부여한다  
용암분출,MONSTER,45,75,15,지면을 뚫고 용암 기둥을 솟아오르게 해 적을 태운다  
파멸의망치,HUMAN,45,80,20,거대한 망치를 휘둘러 적 하나를 강타한다  
혼돈의폭발,GHOST,50,75,5,혼돈의 에너지를 폭발시켜 광역 흡수 피해를 입힌다  
'''

# 두 번째 기술 데이터 (예시로 일부 기술을 수정/추가/삭제)
skill_data_2 = '''
# 초급 기술
가시털,MONSTER,25,85,20,등의 가시를 세워 적을 찔러 공격한다
괴성,GHOST,25,90,25,듣기 싫은 소리를 내 공격한다
그림자발바닥,GHOST,20,100,20,바닥의 그림자를 타고 빠르게 돌진해 공격한다
꼬리치기,ANIMAL,20,100,25,꼬리로 적을 공격한다
날개베기,ANIMAL,25,95,25,날개를 이용해 회전 베기를 가한다
독꽃던지기,EVIL_SPIRIT,15,100,30,독이 깃든 꽃잎을 흩뿌려 중독 상태를 유발한다
독침,ANIMAL,20,100,30,독침을 쏘아 공격한다
돌가루,EVIL_SPIRIT,15,100,30,돌가루를 날려 적의 시야를 가린다
돌던지기,HUMAN,25,95,25,돌을 던져 적을 공격한다
망령의눈빛,GHOST,20,100,30,적에게 오싹한 눈빛을 보내 공격한다
몸통박치기,MONSTER,25,100,20,몸 전체로 부딪혀 공격한다
물기,ANIMAL,25,100,20,이빨로 적을 물어뜯는다
바람발,ANIMAL,15,100,30,날개짓으로 찬 바람을 일으켜 적을 얼린다
부채베기,HUMAN,25,95,20,부채로 예리하게 베어 공격한다
불꽃치기,MONSTER,25,100,20,불꽃으로 적을 공격한다
불씨던지기,MONSTER,25,90,20,작은 불씨를 던져 화상을 입힌다
사슬채찍,ANIMAL,20,100,25,사슬 같은 꼬리로 적을 휘감아 공격한다
새벽의속삭임,EVIL_SPIRIT,15,85,30,새벽녘의 기운으로 적을 혼란에 빠뜨린다
소리지르기,HUMAN,20,100,30,소리를 질러 적을 공격한다
속임수,EVIL_SPIRIT,25,95,25,기습적인 공격을 한다
안개,GHOST,25,100,30,안개에 숨어 적을 공격한다
유혹하기,ANIMAL,15,85,30,적을 유혹해 상대를 혼란스럽게 한다
음파,GHOST,25,90,25,강한 음파로 정신 피해를 입힌다
정전기,GHOST,20,100,30,정전기를 일으켜격한다
진흙던지기,HUMAN,20,100,25,진흙을 뿌려 시야를 방해한다
짖기,ANIMAL,15,100,30,소리로 적을 공격한다
할퀴기,ANIMAL,25,100,20,날카로운 발톱으로 할퀴어 공격한다
헛소리,EVIL_SPIRIT,20,100,30,정신을 혼미하게 만드는 알 수 없는 소리를 낸다
호미찌르기,HUMAN,25,100,20,호미날로 적을 강하게 찌른다

# 중급 기술
가시돌진,MONSTER,40,85,15,몸의 가시로 돌진하며 상대에게 큰 피해를 준다
가시벽,ANIMAL,30,90,20,빠르게 솟아나는 가시로 상대를 가둬 공격한다
검은연기,GHOST,30,90,25,검은 연기를 뿜어 정신을 혼미하게 만든다
검은연못,GHOST,30,85,20,어둠의 연못을 소환해 적을 빨아들인다
괴성공명,GHOST,40,80,15,귀신의 절규가 공명을 일으켜 타격을 준다
귀신의사슬,EVIL_SPIRIT,35,90,15,유령 사슬로 적을 묶어 묵직한 피해를 준다
꼬리후려치기,ANIMAL,35,90,15,날렵한 꼬리로 적을 후려친다
달그림자베기,EVIL_SPIRIT,35,90,15,달빛 속 그림자로 벤다
달빛무도,EVIL_SPIRIT,35,90,15,달빛 아래 환영 무도를 펼쳐 연속 공격을 가한다
마구할퀴기,ANIMAL,40,85,10,연속으로 할퀴어 강한 피해를 준다
망각의노래,EVIL_SPIRIT,35,90,15,저승의 멜로디로 적의 기억을 일시적으로 잃게 한다
망자의손,GHOST,35,85,15,죽은 자의 손이 나타나 덮친다
맹독침,ANIMAL,35,85,20,강한 독을 가진 독침을 쏘아 공격한다
무쇠발톱,ANIMAL,35,85,15,단단한 무쇠 같은 발톱으로 적을 내리찍는다
물어뜯기,ANIMAL,35,85,20,이빨로 적을 물어뜯는다
바람칼,HUMAN,35,85,20,바람으로 만든 칼날을 날린다
번개,MONSTER,40,85,15,번개로 적을 공격한다
불길장벽,MONSTER,35,85,15,불길의 장벽을 세워 적을 태운다
속임수베기,EVIL_SPIRIT,35,90,10,기습적인 공격을 한다
쇠방울던지기,HUMAN,30,90,25,쇠방울을 던져 타격을 준다
숨은송곳니,ANIMAL,35,90,15,숨어 있다가 날카로운 이빨로 덮친다
어퍼컷,HUMAN,35,85,20,강한 어퍼컷을 날펴 피해를 입힌다
얼음창,MONSTER,40,80,15,얼음으로 만든 창을 던진다
연속박치기,MONSTER,40,80,15,연속으로 머리로 들이받는다
염력,GHOST,35,85,20,염력으로 물건을 던져져 공격한다
영혼폭포,GHOST,35,85,15,영혼의 물줄기를 폭포처럼 쏟아내 피해를 준다
예측공격,HUMAN,30,90,30,상대의 움직임을 파악해서 약점을 공격한다
요상한바람,EVIL_SPIRIT,30,90,20,기분 나쁜 바람으로 공격 한다
울부짖기,ANIMAL,30,80,15,큰 소리를 질러 공격한다
철화살,HUMAN,35,85,20,단단한 철제 화살을 날려 정확히 관통한다
칼날비,HUMAN,35,80,20,허공에 나타난 칼날을 비처럼 쏟아낸다
폭음파,GHOST,35,85,25,강한 폭발을 일으켜 만든 음파로 피해를 입힌다
폭풍의깃,MONSTER,40,80,15,회오리치는 폭풍구름을 만들어 적에게 돌진시킨다
피눈물베기,EVIL_SPIRIT,40,85,10,증오로 울며 베어내는 원혼의 칼날
호랑이눈빛,ANIMAL,35,85,20,호랑이의 기운으로 상대를 얼어붙게 한다
호랑이발톱,ANIMAL,30,90,90,호랑이의 발톱으로 강하게 찢어낸다
혼령의숨,GHOST,30,95,20,혼령이 뿜는 숨결로 정신을 흔든다
환영분열,EVIL_SPIRIT,35,85,20,자신의 환영을 나누어 다수 타격을 가한다
환영술,EVIL_SPIRIT,35,85,15,상대의 두려움을 담은 환영 보여준다
후려치기,HUMAN,35,85,20,무거운 일격으로 적을 가격한다
흙폭탄,HUMAN,30,90,20,흙으로 만든 폭탄을 던진다

# 고급 기술
강철꼬리치기,ANIMAL,45,80,10,단단한 꼬리로 강하게 내리친다
그림자베기,GHOST,45,85,10,그림자로 적을 베어버린다
깨물어부수기,ANIMAL,45,85,15,이빨로 깨물어 부순다다
뇌우,GHOST,50,75,5,천둥 번개와 비를 내려 공격한다
달밤유성,GHOST,45,85,10,달빛 속에서 유성이 떨어져 적을 강타한다
도깨비방망이,MONSTER,50,85,5,도깨비방망이로 내려쳐 적을 공격한다
맹화,MONSTER,50,85,5,강한 불꽃으로 공격헌다
무덤기운,GHOST,50,75,5무덤에서 솟아오르는 힘으로 공격한다
무지개발차기,HUMAN,45,80,10,어디서 날라올지 모르는 강력한 발차기를 날린다
백호의울음,ANIMAL,50,75,5,전설 속 백호의 울음으로 적을 압도한다
분노의주먹,HUMAN,45,80,10,분노를 담아 강한 한 방을 날린다
불사조의노래,MONSTER,45,80,10,불사조의 울음으로 전장을 불태우며 부활 효과를 부여한다
사념의박치기,MONSTER,50,80,5,강하게 부딪힌다
암흑회오리,GHOST,50,75,5,암흑 에너지의 회오리를 만들어 광역 흡수 피해를 준다
어둠손길,GHOST,45,85,15,상대를 자신이 있는 어둠속으로 끌어들여 공격한다
영혼흡수,EVIL_SPIRIT,40,80,20,적의 영혼을 흡수한다
용암분출,MONSTER,45,75,15,지면을 뚫고 용암 기둥을 솟아오르게 해 적을 태운다
이단도약,MONSTER,45,85,10,두 번 튕겨 뛰며 큰 피해를 준다
지옥기운,EVIL_SPIRIT,50,70,5,지옥의 어둠으로 적에게 큰 피해를 준다
지옥기운,EVIL_SPIRIT,50,70,5,지옥의 어둠으로 적에게 큰 피해를 준다
지옥불,EVIL_SPIRIT,50,70,5,지옥의 불꽃으로 적을 태운다
지옥폭뢰,EVIL_SPIRIT,50,70,10,지옥의 번개를 소환해 전장을 가로지른다
천지개벽,MONSTER,50,75,5,하늘과 땅을 가르는 거대한 충격파를 발사한다
철퇴내리치기,HUMAN,50,75,5,무거운 철퇴를 휘둘러 큰 피해를 준다
태산벽력,MONSTER,50,75,5,산을 깨뜨리는 벽력으로 일격을 가한다
파멸의망치,HUMAN,45,80,20,거대한 망치를 휘둘러 적 하나를 강타한다
혼돈의폭발,GHOST,50,75,5,혼돈의 에너지를 폭발시켜 광역 흡수 피해를 입힌다
혼령메아리,GHOST,45,80,10,죽은 혼령들의 메아리로 정신 공격을 가한다
혼백폭발,GHOST,50,75,5,혼백의 힘으로 폭발을 일으킨다
혼신의부채베기,HUMAN,50,80,5,혼신을 담아 부채로 빠르게 베어낸다

'''

def parse_skill_data(text_data):
    """텍스트 형식의 기술 데이터를 파싱합니다."""
    skill_records = []
    current_difficulty = "초급"  # 기본 난이도
    
    for line in text_data.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith('#'):
            if "초급" in line:
                current_difficulty = "초급"
            elif "중급" in line:
                current_difficulty = "중급"
            elif "고급" in line:
                current_difficulty = "고급"
            elif "패러독스" in line:
                current_difficulty = "패러독스"
            elif "최종보스" in line:
                current_difficulty = "최종보스"
            continue
        
        # 주석이나 형식 설명 라인은 건너뛰기
        if "형식:" in line or "기술 데이터 파일" in line:
            continue
        
        parts = line.split(",")
        if len(parts) >= 6:
            name = parts[0].strip()
            type_ = parts[1].strip()
            power = parts[2].strip()
            accuracy = parts[3].strip()
            pp = parts[4].strip()
            description = parts[5].strip()
            
            # PP 값이 '/'로 시작하는 경우 처리
            if pp.startswith('/'):
                pp = pp[1:]  # '/' 제거
            
            skill_records.append([name, type_, power, accuracy, pp, description, current_difficulty])
    
    return pd.DataFrame(skill_records, columns=["이름", "타입", "위력", "명중률", "PP", "설명", "난이도"])

def compare_skills(df1, df2):
    """두 기술 모음을 비교하여 차이점을 찾습니다."""
    # 이름을 기준으로 병합
    merged = pd.merge(df1, df2, on='이름', how='outer', suffixes=('_1', '_2'))
    
    # 차이점이 있는 행만 필터링
    differences = []
    for _, row in merged.iterrows():
        # 첫 번째 데이터에만 있는 기술 (삭제됨)
        if pd.isna(row['타입_2']):
            diff = {
                '이름': row['이름'],
                '상태': '삭제됨',
                '타입_1': row['타입_1'],
                '타입_2': None,
                '위력_1': row['위력_1'],
                '위력_2': None,
                '명중률_1': row['명중률_1'],
                '명중률_2': None,
                'PP_1': row['PP_1'],
                'PP_2': None,
                '설명_1': row['설명_1'],
                '설명_2': None,
                '난이도_1': row['난이도_1'],
                '난이도_2': None
            }
            differences.append(diff)
        # 두 번째 데이터에만 있는 기술 (추가됨)
        elif pd.isna(row['타입_1']):
            diff = {
                '이름': row['이름'],
                '상태': '추가됨',
                '타입_1': None,
                '타입_2': row['타입_2'],
                '위력_1': None,
                '위력_2': row['위력_2'],
                '명중률_1': None,
                '명중률_2': row['명중률_2'],
                'PP_1': None,
                'PP_2': row['PP_2'],
                '설명_1': None,
                '설명_2': row['설명_2'],
                '난이도_1': None,
                '난이도_2': row['난이도_2']
            }
            differences.append(diff)
        # 두 데이터 모두에 있는 기술 (변경됨)
        else:
            if (row['타입_1'] != row['타입_2'] or
                row['위력_1'] != row['위력_2'] or
                row['명중률_1'] != row['명중률_2'] or
                row['PP_1'] != row['PP_2'] or
                row['설명_1'] != row['설명_2'] or
                row['난이도_1'] != row['난이도_2']):
                diff = {
                    '이름': row['이름'],
                    '상태': '변경됨',
                    '타입_1': row['타입_1'],
                    '타입_2': row['타입_2'],
                    '위력_1': row['위력_1'],
                    '위력_2': row['위력_2'],
                    '명중률_1': row['명중률_1'],
                    '명중률_2': row['명중률_2'],
                    'PP_1': row['PP_1'],
                    'PP_2': row['PP_2'],
                    '설명_1': row['설명_1'],
                    '설명_2': row['설명_2'],
                    '난이도_1': row['난이도_1'],
                    '난이도_2': row['난이도_2']
                }
                differences.append(diff)
    
    return pd.DataFrame(differences)

def save_to_excel(df, output_path):
    """비교 결과를 엑셀 파일로 저장합니다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "기술 비교 결과"
    
    # 데이터 입력
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 상태별 색상 지정
    status_colors = {
        '추가됨': '99FF99',  # 연한 초록색
        '삭제됨': 'FF9999',  # 연한 빨간색
        '변경됨': 'FFFF99'   # 연한 노란색
    }
    
    # 상태 열에 색상 적용
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            color = status_colors.get(cell.value, None)
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # 표 추가
    table = Table(displayName="SkillComparison", ref=f"A1:N{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # 저장
    wb.save(output_path)
    print(f"비교 결과가 {output_path}에 저장되었습니다.")

def main():
    # 데이터 파싱
    df1 = parse_skill_data(skill_data_1)
    df2 = parse_skill_data(skill_data_2)
    
    # 비교
    differences = compare_skills(df1, df2)
    
    # 결과 저장
    output_path = "기술_비교_결과.xlsx"
    save_to_excel(differences, output_path)

if __name__ == "__main__":
    main() 