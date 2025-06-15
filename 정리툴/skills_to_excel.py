import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

# 기술 데이터 (여기에 붙여넣거나, 외부 파일에서 읽어올 수도 있음)
skill_data = '''
# 기술 데이터 파일
# 형식: 이름,기술의 상성,공격력,명중률,PP,설명

# 초급 기술
불꽃치기,MONSTER,30,90,20,불꽃으로 적을 공격한다
돌던지기,HUMAN,25,95,25,돌을 던져 적을 공격한다
꼬리치기,ANIMAL,20,100,25,꼬리로 적을 공격한다
물기,ANIMAL,25,90,20,이빨로 적을 물어뜯는다
할퀴기,ANIMAL,25,90,20,날카로운 발톱으로 할퀴어 공격한다
몸통박치기,MONSTER,30,85,20,몸 전체로 부딪혀 공격한다
독침,ANIMAL,20,95,30,독침을 쏘아 공격한다
짖기,ANIMAL,15,100,30,큰 소리로 적의 명중률을 낮춘다
음파,GHOST,25,90,25,강한 음파로 정신 피해를 입힌다
유혹하기,ANIMAL,15,85,30,적을 유혹하여 공격력을 낮춘다
날개베기,ANIMAL,30,90,15,날개를 이용해 회전 베기를 가한다
속임수베기,HUMAN,35,100,10,기습 공격으로반드시 먼저 공격한다
박치기,MONSTER,30,95,15,연속으로 머리로 들이받는다

# 중급 기술
바람칼,HUMAN,35,85,20,바람으로 만든 칼날을 날린다
얼음창,GHOST,40,80,15,얼음으로 만든 창을 던진다
물어뜯기,ANIMAL,35,80,20,이빨로 적을 물어뜯는다
후려치기,HUMAN,35,80,20,무거운 일격으로 적을 가격한다
맹독침,ANIMAL,35,80,20,강한 독을 가진 독침을 쏘아 공격한다
흙폭탄,HUMAN,30,90,20,흙으로 만든 폭탄을 던진다
마구 할퀴기,ANIMAL,40,85,10,연속으로 할퀴어 강한 피해를 준다
염력,GHOST,35,85,20,염력으로 물건을 던져져 공격한다
울부짖기,ANIMAL,20,100,5,자신의 공격력을 증가시킨다
연속박치기,MONSTER,40,80,/15,연속으로 머리로 들이받는다

# 고급 기술
지옥불,EVIL_SPIRIT,50,70,5,지옥의 불꽃으로 적을 태운다
가시돌진,MONSTER,40,85,15,몸의 가시로 돌진하며 상대에게 큰 피해를 준다
분노의주먹,HUMAN,45,80,10,분노를 담아 강한 한 방을 날린다
사념의박치기,MONSTER,50,80,5,강하게 부딪힌다다
강철꼬리치기,MONSTER,45,80,10,단단한 꼬리로 강하게 내리친다
맹화,MONSTER,50,85,5,강한 불꽃으로 공격헌다
깨물어부수기,ANIMAL,40,90,15,이빨로 깨물어 부순다다
번개,GHOST,45,75,10,번개로 적을 공격한다
철퇴내리치기,HUMAN,50,75,5,무거운 철퇴를 휘둘러 큰 피해를 준다
지옥기운,EVIL_SPIRIT,60,65,5,지옥의 어둠으로 적에게 큰 피해를 준다
이단도약,MONSTER,45,85,10,두 번 튕겨 뛰며 큰 피해를 준다
망령의눈빛,GHOST,100,30,5,행동불능으로 만들 확률 30%

# 패러독스 요괴 기술
혼돈의파동,EVIL_SPIRIT,85,90,10,혼돈의 파동으로 적을 공격
차원의칼날,EVIL_SPIRIT,80,95,10,차원의 칼날로 적을 베어버림
어둠의손길,EVIL_SPIRIT,75,100,15,어둠의 손길로 적을 공격
영혼흡수,EVIL_SPIRIT,70,100,15,적의 영혼을 흡수하여 HP 회복
차원방패,EVIL_SPIRIT,0,100,10,차원의 방패로 자신을 보호
파멸의주먹,EVIL_SPIRIT,85,90,10,파멸을 부르는 강력한 주먹
공간왜곡,EVIL_SPIRIT,0,100,5,공간을 왜곡하여 적의 공격을 무효화
수호의외침,EVIL_SPIRIT,0,100,5,강력한 외침으로 자신의 방어력 증가
그림자베기,GHOST,75,95,10,그림자로 적을 베어버림
암흑폭풍,GHOST,80,90,10,암흑의 폭풍으로 적을 공격
속삭임,GHOST,0,100,5,적의 공격력을 낮추는 속삭임
절망의눈빛,GHOST,0,100,5,적을 절망에 빠뜨리는 눈빛
운명베기,HUMAN,80,90,10,운명을 베어버리는 강력한 공격
예지타격,HUMAN,75,95,10,미래를 예지하여 정확한 타격
시간왜곡,HUMAN,0,100,5,시간을 왜곡하여 적의 행동을 봉쇄
심판의칼날,HUMAN,85,85,10,심판의 칼날로 적을 베어버림
시간단절,EVIL_SPIRIT,0,100,5,시간의 흐름을 끊어버림
과거회상,EVIL_SPIRIT,0,100,5,과거의 기억으로 자신을 강화
미래도약,EVIL_SPIRIT,0,100,5,미래로 도약하여 회피
영원의속삭임,EVIL_SPIRIT,0,100,5,영원한 속삭임으로 적을 약화
공허의창,EVIL_SPIRIT,80,90,10,공허의 창으로 적을 관통
무감정공격,EVIL_SPIRIT,75,95,10,감정 없는 공격으로 적을 공격
침묵의포효,EVIL_SPIRIT,0,100,5,침묵의 포효로 적을 약화
무의의손길,EVIL_SPIRIT,70,100,15,무의의 손길로 적을 공격
붕괴의손짓,GHOST,0,100,5,붕괴의 손짓으로 적을 약화
파멸의속삭임,GHOST,0,100,5,파멸의 속삭임으로 적을 공격
무의식공격,GHOST,75,95,10,무의식의 공격으로 적을 공격
종말의울음,GHOST,0,100,5,종말을 알리는 울음으로 적을 약화
경계이동,HUMAN,0,100,5,경계를 이동하여 회피
방랑자의칼날,HUMAN,80,90,10,방랑자의 칼날로 적을 베어버림
이방인의눈,HUMAN,0,100,5,이방인의 눈으로 적을 약화
자유의외침,HUMAN,0,100,5,자유의 외침으로 자신을 강화
환영공격,ANIMAL,75,95,10,환영으로 적을 공격
실체변화,ANIMAL,0,100,5,실체를 변화시켜 회피
환각의포효,ANIMAL,0,100,5,환각의 포효로 적을 혼란
허상의손길,ANIMAL,70,100,15,허상의 손길로 적을 공격

# 최종보스 기술
차원파괴,EVIL_SPIRIT,100,80,5,차원을 파괴하는 강력한 공격
시간지배,EVIL_SPIRIT,90,85,5,시간을 지배하여 적의 행동을 봉쇄
공간지배,EVIL_SPIRIT,95,80,5,공간을 지배하여 적을 압박
'''

# 타입별 색상 지정
type_colors = {
    "EVIL_SPIRIT": "FF9999",  # 빨간색
    "GHOST": "CC99FF",        # 보라색
    "MONSTER": "FFFF99",      # 노란색
    "HUMAN": "66CCCC",        # 청록색
    "ANIMAL": "99CC99",       # 초록색
}

# 난이도별 색상 지정
difficulty_colors = {
    "초급": "FFFFFF",  # 흰색
    "중급": "FFCC99",  # 연한 주황색
    "고급": "FF9999",  # 연한 빨간색
    "패러독스": "CC99FF",  # 보라색
    "최종보스": "FF6666",  # 진한 빨간색
}

# 기술 데이터 파싱 (주석, 공백 줄 무시)
skill_records = []
current_difficulty = "초급"  # 기본 난이도

for line in skill_data.strip().splitlines():
    line = line.strip()
    if not line:
        continue
    if line.startswith('#'):
        if "초급" in line:
            current_difficulty = "초급"
        elif "중급" in line:
            current_difficulty = "중급"
        elif "고급" in line:
            current_difficulty = "고급"
        elif "패러독스" in line:
            current_difficulty = "패러독스"
        elif "최종보스" in line:
            current_difficulty = "최종보스"
        continue
    
    parts = line.split(",")
    if len(parts) >= 6:
        name = parts[0].strip()
        type_ = parts[1].strip()
        power = parts[2].strip()
        accuracy = parts[3].strip()
        pp = parts[4].strip()
        description = parts[5].strip()
        skill_records.append([name, type_, power, accuracy, pp, description, current_difficulty])

# DataFrame 생성
df = pd.DataFrame(skill_records, columns=["이름", "타입", "위력", "명중률", "PP", "설명", "난이도"])

# 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "기술 도감"

# 데이터 입력
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 타입별 색상 적용
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    for cell in row:
        color = type_colors.get(cell.value, None)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# 난이도별 색상 적용
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
    for cell in row:
        color = difficulty_colors.get(cell.value, None)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# 표 추가
table = Table(displayName="SkillDex", ref=f"A1:G{ws.max_row}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws.add_table(table)

# 열 너비 자동 조정
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# 저장
path = "기술_도감.xlsx"
wb.save(path)

print(f"기술 도감이 {path}에 저장되었습니다.") 